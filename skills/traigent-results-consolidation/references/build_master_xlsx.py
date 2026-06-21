"""Build the aggregate master workbook — one row per eval, ALL runs — per the
`traigent-results-consolidation` column contract. Reads results/*.jsonl (written
by run.py after each real run) and emits a sortable .xlsx for cross-run analysis.

  python build_master_xlsx.py     # -> results/SPIDER_text2sql_ALL_runs_master.xlsx

Columns (left->right): run · permutations · seq · weighted_score · accuracy ·
cost_usd · latency_s · model (full+decomposed) · all knobs (union, first-
appearance order). Header auto-filter + frozen panes; a second "Runs" sheet legends
each batch. Requires openpyxl (`pip install openpyxl`).
"""

from __future__ import annotations

import glob
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
RESULTS = ROOT / "results"
META = {"run", "permutations", "weights", "accuracy", "cost", "latency"}


def load_rows() -> list[dict]:
    rows = []
    for fp in sorted(glob.glob(str(RESULTS / "*.jsonl"))):
        for line in Path(fp).read_text(encoding="utf-8").splitlines():
            if line.strip():
                rows.append(json.loads(line))
    return rows


def decompose_model(m: str) -> tuple[str, str, str]:
    """openrouter/openai/gpt-4o-mini -> (name, number, vendor)."""
    parts = str(m).split("/")
    vendor = parts[1] if len(parts) >= 3 else (parts[0] if parts else "")
    name = parts[-1] if parts else ""
    num = re.search(r"[0-9][0-9.]*[a-z-]*", name)
    return name, (num.group(0) if num else ""), vendor


def benefit(values: list[float]) -> dict[float, float]:
    """Min-max 'benefit' for a MINIMIZED metric: lowest -> 1.0, highest -> 0.0.
    Constant/absent -> 1.0 (neutral)."""
    lo, hi = min(values), max(values)
    if hi == lo:
        return {v: 1.0 for v in values}
    return {v: (hi - v) / (hi - lo) for v in values}


def main() -> int:
    rows = load_rows()
    if not rows:
        print("No results/*.jsonl yet — complete a real run first.")
        return 0

    # knob columns = union across rows, first-appearance order (excl. meta + model)
    knob_cols: list[str] = []
    for r in rows:
        for k in r:
            if k not in META and k != "model" and k not in knob_cols:
                knob_cols.append(k)

    # batch letters A,B,C... by run chronological (first-appearance) order
    runs: list[str] = []
    for r in rows:
        if r["run"] not in runs:
            runs.append(r["run"])
    letter = {run: chr(ord("A") + i) for i, run in enumerate(runs)}

    # per-run min-max benefit for the minimized metrics (for weighted_score)
    per_run = defaultdict(list)
    for r in rows:
        per_run[r["run"]].append(r)
    cost_ben = {run: benefit([x["cost"] for x in rs]) for run, rs in per_run.items()}
    lat_ben = {run: benefit([x["latency"] for x in rs]) for run, rs in per_run.items()}

    wb = Workbook()
    ws = wb.active
    ws.title = "All evals"
    headers = ["run", "permutations", "seq", "weighted_score", "accuracy",
               "cost_usd", "latency_s", "model_full", "model_name",
               "model_number", "model_vendor"] + knob_cols
    ws.append(headers)

    seqn = defaultdict(int)
    for r in rows:
        run = r["run"]
        seqn[run] += 1
        seq = f"{letter[run]}{seqn[run]}"
        w = r.get("weights", {}) or {}
        score = (w.get("accuracy", 0) * r["accuracy"]
                 + w.get("cost", 0) * cost_ben[run][r["cost"]]
                 + w.get("latency", 0) * lat_ben[run][r["latency"]])
        name, num, vendor = decompose_model(r.get("model", ""))
        ws.append([run, r.get("permutations"), seq, round(score, 4), r["accuracy"],
                   r["cost"], r["latency"], r.get("model"), name, num, vendor]
                  + [r.get(k, "") for k in knob_cols])

    hdr_fill = PatternFill("solid", fgColor="1F2747")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # Runs legend sheet
    rs = wb.create_sheet("Runs")
    rs.append(["batch", "run", "permutations", "weights", "#evals"])
    for run in runs:
        rws = per_run[run]
        rs.append([letter[run], run, rws[0].get("permutations"),
                   json.dumps(rws[0].get("weights", {})), len(rws)])
    for c in range(1, 6):
        rs.cell(1, c).font = Font(bold=True)

    out = RESULTS / "SPIDER_text2sql_ALL_runs_master.xlsx"
    wb.save(out)
    print(f"wrote {out}  ({len(rows)} eval rows across {len(runs)} runs; "
          f"weighted_score = w_acc*acc + w_cost*cost_benefit + w_lat*lat_benefit)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
